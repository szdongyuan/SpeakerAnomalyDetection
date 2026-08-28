import numpy as np
import pytest
from openpyxl import load_workbook

from base.excel_result_exporter import (
    _export_unit,
    _extract_curve_xy,
    _iter_selected_runtime_items,
    _make_margin_sheet_title,
    _sanitize_sheet_name,
    build_excel_from_csv_spool,
    export_analysis_to_csv_spool,
    export_analysis_to_excel,
)


def test_extract_curve_xy_prefers_raw_keys_when_present():
    result = {
        "freq_value": [1.0, 2.0, 3.0],
        "thd": [10.0, 11.0, 12.0],
        "thd_raw": [100.0, 110.0, 120.0],
    }
    x, y = _extract_curve_xy(result)
    assert x == [1.0, 2.0, 3.0]
    assert y == [100.0, 110.0, 120.0]


def test_extract_curve_xy_falls_back_to_display_keys_when_raw_missing():
    result = {"freq_value": [1.0, 2.0], "thd": [10.0, 11.0]}
    x, y = _extract_curve_xy(result)
    assert x == [1.0, 2.0]
    assert y == [10.0, 11.0]


def test_extract_curve_xy_handles_numpy_arrays_and_prefers_raw():
    result = {
        "frequency_list": np.asarray([100.0, 200.0], dtype=float),
        "fr": np.asarray([-1.0, -2.0], dtype=float),
        "fr_raw": np.asarray([-10.0, -20.0], dtype=float),
    }
    x, y = _extract_curve_xy(result)
    assert x == [100.0, 200.0]
    assert y == [-10.0, -20.0]


def test_extract_curve_xy_supports_fba_weighted_band_levels():
    result = {
        "band_centers": [100.0, 1000.0],
        "band_levels_db": [60.0, 70.0],
        "band_levels_weighted_db": [41.0, 70.0],
    }

    x, y = _extract_curve_xy(result)

    assert x == [100.0, 1000.0]
    assert y == [41.0, 70.0]
    assert _export_unit("FBA", {}) == "dB"


def test_extract_curve_xy_supports_fft_display_curve():
    result = {
        "frequency_bins": [100.0, 1000.0],
        "fft_db": [60.0, 70.0],
        "plot_db": [-2.0, 3.0],
    }

    x, y = _extract_curve_xy(result)

    assert x == [100.0, 1000.0]
    assert y == [-2.0, 3.0]
    assert _export_unit("FFT", {}) == "dB"


def test_extract_curve_xy_supports_loudness_curve():
    result = {
        "time_s": [0.05, 0.15],
        "loudness_sone": [1.0, 1.5],
    }

    x, y = _extract_curve_xy(result)

    assert x == [0.05, 0.15]
    assert y == [1.0, 1.5]
    assert _export_unit("LOUD", {}) == "sone"


@pytest.mark.parametrize("expanded", [False, True])
def test_single_export_keeps_legacy_name_unless_other_selected_channels_are_missing(expanded):
    item = {
        "config_key": "SPL 1",
        "result_key": "SPL 1--通道2",
        "multi_channel_expansion": expanded,
        "type": "SPL",
    }

    selected = list(_iter_selected_runtime_items(
        ["SPL 1"], {"SPL 1--通道2": item}
    ))

    expected_name = "SPL 1--通道2" if expanded else "SPL 1"
    assert selected == [(expected_name, "SPL 1", "SPL 1--通道2", item)]


@pytest.mark.parametrize("use_spool", [False, True])
def test_excel_export_writes_runtime_channels_to_separate_sheets(tmp_path, use_spool):
    excel_cfg = {
        "enabled": True,
        "save_items": ["SPL 1"],
        "save_dir": str(tmp_path),
        "file_base": "multi_channel",
        "add_date": False,
        "lock_files": False,
    }
    data = {
        f"SPL 1--通道{channel}": {
            "config_key": "SPL 1",
            "result_key": f"SPL 1--通道{channel}",
            "type": "SPL",
            "multi_channel_expansion": True,
            "result": {
                "signal_duration": [0.0, 0.1],
                "signal_spl": [value, value + 1],
            },
        }
        for channel, value in [(1, 60.0), (3, 70.0)]
    }
    export = export_analysis_to_csv_spool if use_spool else export_analysis_to_excel
    result = export(
        excel_cfg,
        sn="SN001",
        date_text="2026/8/27 18:00:00",
        analysis_items_data=data,
        analysis_config={"SPL 1": {"type": "SPL"}},
        analysis_result_dict={
            "SPL 1--通道1": (True, 0.0),
            "SPL 1--通道3": (False, 1.0),
        },
    )

    assert result.ok is True
    if use_spool:
        assert build_excel_from_csv_spool(excel_cfg).ok is True
    workbook = load_workbook(tmp_path / "multi_channel.xlsx")
    try:
        assert "SPL 1--通道1" in workbook.sheetnames
        assert "SPL 1--通道3" in workbook.sheetnames
        assert workbook["SPL 1--通道1"].cell(row=2, column=3).value == 60.0
        assert workbook["SPL 1--通道3"].cell(row=2, column=3).value == 70.0
    finally:
        workbook.close()


def test_long_excel_names_keep_channel_suffix_unique():
    base_name = "这是一个非常长的声压级分析项目名称用于验证Excel限制"
    first = _sanitize_sheet_name(f"{base_name}--通道1")
    third = _sanitize_sheet_name(f"{base_name}--通道3")
    margin_first = _make_margin_sheet_title(f"{base_name}--通道1")
    margin_third = _make_margin_sheet_title(f"{base_name}--通道3")

    assert max(map(len, [first, third, margin_first, margin_third])) <= 31
    assert first.endswith("--通道1")
    assert third.endswith("--通道3")
    assert first != third
    assert margin_first.endswith("--通道1 margin")
    assert margin_third.endswith("--通道3 margin")
    assert margin_first != margin_third
