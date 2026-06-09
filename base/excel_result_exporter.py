from __future__ import annotations

import io
import os
import re
import csv
import threading
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Iterable

import numpy as np
from openpyxl import Workbook, load_workbook
import msvcrt  # Windows-only; used for runtime file locking

from consts.running_consts import DEFAULT_DIR





EXCEL_MAX_COLUMNS = 16384  # hard limit
EXCEL_RESERVED_COLUMNS = 2  # SN, 日期
EXCEL_MAX_DATA_POINTS = EXCEL_MAX_COLUMNS - EXCEL_RESERVED_COLUMNS


_INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]\:\*\?\/\\]")
_INVALID_FILENAME_CHARS_RE = re.compile(r"[<>:\"/\\\\|?*]")


@dataclass(frozen=True)
class ExportResult:
    ok: bool
    message: str


_RUNTIME_FILE_LOCK_BYTES = 1


def _locks_enabled(excel_cfg: dict[str, Any]) -> bool:
    # Default True: prevent operator opening CSV/XLSX while program is running.
    return bool((excel_cfg or {}).get("lock_files", True))


class _ExportLockError(Exception):
    pass


@dataclass
class _LockedTextFile:
    path: str
    f: Any  # TextIO
    lock: threading.Lock
    writer: csv.writer
    header: list[Any] | None = None


@dataclass
class _LockedBinaryFile:
    path: str
    f: Any  # BinaryIO
    lock: threading.Lock


class _RuntimeFileLocker:
    """
    Hold mandatory locks (Windows) to prevent other processes (Excel) from opening
    CSV spool files and daily XLSX while the program is running.

    Notes:
    - Uses a 1-byte lock at offset 0. This is enough to block typical readers which
      must read the header from the beginning of the file.
    - While locked, *this process* must reuse the same file handle; opening the
      same path again will raise PermissionError on Windows.
    """

    def __init__(self):
        self._mu = threading.Lock()
        self._csv: dict[str, _LockedTextFile] = {}
        self._xlsx: dict[str, _LockedBinaryFile] = {}

    def _headers_equivalent(self, existing: list[Any], expected: list[Any]) -> bool:
        if len(existing) != len(expected):
            return False
        a = [_normalize_header_value(_coerce_csv_number(v)) for v in existing]
        b = [_normalize_header_value(_coerce_csv_number(v)) for v in expected]
        return _headers_match(a, b)

    def _lock_file(self, f) -> None:
        if msvcrt is None:
            return
        f.seek(0)
        msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, _RUNTIME_FILE_LOCK_BYTES)

    def _unlock_file(self, f) -> None:
        if msvcrt is None:
            return
        f.seek(0)
        msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, _RUNTIME_FILE_LOCK_BYTES)

    def append_csv_row(self, path: str, *, header: list[Any], row: list[Any]) -> ExportResult:
        path = str(path)
        try:
            lf = self._get_or_open_csv(path, expected_header=header)
        except _ExportLockError as e:
            return ExportResult(ok=False, message=str(e))
        except PermissionError:
            return ExportResult(ok=False, message=f"CSV文件被占用，请关闭后重试: {path}")
        except Exception as e:
            return ExportResult(ok=False, message=f"CSV加锁失败: {path}: {e}")

        try:
            with lf.lock:
                lf.writer.writerow(row)
                try:
                    lf.f.flush()
                except Exception:
                    pass
            return ExportResult(ok=True, message=f"CSV追加成功: {path}")
        except PermissionError:
            return ExportResult(ok=False, message=f"CSV文件被占用，请关闭后重试: {path}")
        except Exception as e:
            return ExportResult(ok=False, message=f"写入CSV失败: {e}")

    def read_csv_snapshot(self, path: str) -> tuple[ExportResult, list[Any] | None, str]:
        """Return (result, header, content) where content is a full text snapshot."""
        path = str(path)
        try:
            lf = self._get_or_open_csv(path, expected_header=None)
        except _ExportLockError as e:
            return ExportResult(ok=False, message=str(e)), None, ""
        except PermissionError:
            return ExportResult(ok=False, message=f"CSV文件被占用，请关闭后重试: {path}"), None, ""
        except Exception as e:
            return ExportResult(ok=False, message=f"CSV加锁失败: {path}: {e}"), None, ""

        try:
            with lf.lock:
                try:
                    lf.f.flush()
                except Exception:
                    pass
                lf.f.seek(0)
                content = lf.f.read()
                lf.f.seek(0, os.SEEK_END)
            header = None
            try:
                reader = csv.reader(io.StringIO(content))
                header = next(reader, None)
            except Exception:
                header = None
            return ExportResult(ok=True, message="ok"), (list(header) if header else None), content
        except Exception as e:
            return ExportResult(ok=False, message=f"读取CSV失败: {path}: {e}"), None, ""

    def peek_csv_header(self, path: str) -> list[Any] | None:
        """
        Read the first CSV row using an already-locked handle if available.
        Returns None when the file isn't tracked/locked by this process.
        """
        path = str(path)
        with self._mu:
            lf = self._csv.get(path)
        if lf is None:
            return None
        try:
            with lf.lock:
                try:
                    lf.f.flush()
                except Exception:
                    pass
                lf.f.seek(0)
                reader = csv.reader(lf.f)
                header = next(reader, None)
                lf.f.seek(0, os.SEEK_END)
                return list(header) if header else None
        except Exception:
            return None

    def get_locked_xlsx(self, path: str) -> _LockedBinaryFile:
        path = str(path)
        with self._mu:
            lf = self._xlsx.get(path)
            if lf is not None:
                return lf

            os.makedirs(os.path.dirname(path), exist_ok=True)
            # Important: do NOT use append mode for zip writing (xlsx); zipfile needs seek+overwrite.
            # Don't truncate here; caller will seek/truncate before saving.
            f = open(path, "r+b") if os.path.exists(path) else open(path, "w+b")
            try:
                self._lock_file(f)
            except Exception:
                try:
                    f.close()
                except Exception:
                    pass
                raise
            lf = _LockedBinaryFile(path=path, f=f, lock=threading.Lock())
            self._xlsx[path] = lf
            return lf

    def save_xlsx(self, path: str, wb: Workbook) -> ExportResult:
        try:
            lf = self.get_locked_xlsx(path)
        except PermissionError:
            return ExportResult(ok=False, message=f"Excel文件被占用，请关闭后重试: {path}")
        except Exception as e:
            return ExportResult(ok=False, message=f"Excel加锁失败: {e}")

        try:
            with lf.lock:
                lf.f.seek(0)
                lf.f.truncate(0)
                wb.save(lf.f)
                try:
                    lf.f.flush()
                except Exception:
                    pass
            return ExportResult(ok=True, message=f"Excel生成成功: {path}")
        except PermissionError:
            return ExportResult(ok=False, message=f"Excel文件被占用，请关闭后重试: {path}")
        except Exception as e:
            return ExportResult(ok=False, message=f"保存Excel失败: {e}")

    def close_all(self) -> None:
        with self._mu:
            csv_files = list(self._csv.values())
            xlsx_files = list(self._xlsx.values())
            self._csv.clear()
            self._xlsx.clear()

        for lf in csv_files:
            try:
                with lf.lock:
                    self._unlock_file(lf.f)
                    lf.f.close()
            except Exception:
                pass
        for lf in xlsx_files:
            try:
                with lf.lock:
                    self._unlock_file(lf.f)
                    lf.f.close()
            except Exception:
                pass

    def close_xlsx(self, path: str) -> None:
        path_norm = os.path.normcase(os.path.abspath(str(path)))
        with self._mu:
            match_key = None
            for k in list(self._xlsx.keys()):
                if os.path.normcase(os.path.abspath(str(k))) == path_norm:
                    match_key = k
                    break
            lf = self._xlsx.pop(match_key, None) if match_key is not None else None
        if lf is None:
            return
        try:
            with lf.lock:
                self._unlock_file(lf.f)
                lf.f.close()
        except Exception:
            pass

    def close_csv_under_dir(self, dir_path: str) -> None:
        dir_norm = os.path.normcase(os.path.abspath(str(dir_path)))
        to_close: list[_LockedTextFile] = []
        with self._mu:
            for k, lf in list(self._csv.items()):
                try:
                    k_norm = os.path.normcase(os.path.abspath(str(k)))
                except Exception:
                    continue
                if k_norm.startswith(dir_norm + os.sep):
                    to_close.append(lf)
                    self._csv.pop(k, None)

        for lf in to_close:
            try:
                with lf.lock:
                    self._unlock_file(lf.f)
                    lf.f.close()
            except Exception:
                pass

    def _get_or_open_csv(self, path: str, *, expected_header: list[Any] | None) -> _LockedTextFile:
        with self._mu:
            lf = self._csv.get(path)
            if lf is not None:
                if (
                    expected_header is not None
                    and lf.header is not None
                    and not self._headers_equivalent(list(lf.header), list(expected_header))
                ):
                    raise _ExportLockError(f"CSV表头不一致，请更换保存目录或清理缓存: {path}")
                return lf

            os.makedirs(os.path.dirname(path), exist_ok=True)
            # a+ : allow both append writes and snapshot reads using the same handle.
            f = open(path, "a+", newline="", encoding="utf-8-sig")
            try:
                self._lock_file(f)

                # Ensure header exists and matches (if expected provided).
                writer = csv.writer(f)
                file_size = 0
                try:
                    f.seek(0, os.SEEK_END)
                    file_size = int(f.tell())
                except Exception:
                    file_size = 0

                header_in_file: list[Any] | None = None
                if file_size > 0:
                    try:
                        f.seek(0)
                        reader = csv.reader(f)
                        header_in_file = next(reader, None)
                    except Exception:
                        header_in_file = None
                    finally:
                        try:
                            f.seek(0, os.SEEK_END)
                        except Exception:
                            pass

                if expected_header is not None:
                    if file_size == 0:
                        writer.writerow(list(expected_header))
                    else:
                        if header_in_file is None:
                            raise _ExportLockError(f"CSV表头读取失败: {path}")
                        if not self._headers_equivalent(list(header_in_file), list(expected_header)):
                            raise _ExportLockError(f"CSV表头不一致，请更换保存目录或清理缓存: {path}")

                lf = _LockedTextFile(
                    path=path,
                    f=f,
                    lock=threading.Lock(),
                    writer=writer,
                    header=list(expected_header)
                    if expected_header is not None
                    else (list(header_in_file) if header_in_file else None),
                )
                self._csv[path] = lf
                return lf
            except Exception:
                try:
                    self._unlock_file(f)
                except Exception:
                    pass
                try:
                    f.close()
                except Exception:
                    pass
                raise


_RUNTIME_LOCKER = _RuntimeFileLocker()


def _resolve_product_model_dir_name(product_model: Any) -> str:
    text = "" if product_model is None else str(product_model)
    text = text.strip()
    return text or "空型号"


def resolve_excel_output_path(excel_cfg: dict[str, Any], *, product_model: str | None = None) -> str:
    """
    Resolve Excel output file path from config, creating the target directory if needed.
    """
    save_dir = excel_cfg.get("save_dir") if isinstance(excel_cfg, dict) else None
    if not save_dir:
        save_dir = os.path.join(DEFAULT_DIR, "audio_data", "excel_export")

    # Optional: append product model as a subdirectory under save_dir.
    if bool((excel_cfg or {}).get("add_model_dir", False)):
        model_dir_name = _resolve_product_model_dir_name(product_model)
        save_dir = os.path.join(save_dir, model_dir_name)
    os.makedirs(save_dir, exist_ok=True)

    file_base = _sanitize_filename((excel_cfg or {}).get("file_base") or "analysis_results")
    add_date = bool((excel_cfg or {}).get("add_date", True))
    date_format = (excel_cfg or {}).get("date_format") or "%Y%m%d"
    if add_date:
        suffix = datetime.now().strftime(date_format)
        filename = f"{file_base}_{suffix}.xlsx"
    else:
        filename = f"{file_base}.xlsx"
    return os.path.join(save_dir, filename)


def resolve_excel_max_points(excel_cfg: dict[str, Any]) -> int:
    max_points = int((excel_cfg or {}).get("max_points", 2000) or 2000)
    return max(10, min(max_points, EXCEL_MAX_DATA_POINTS))


class ExcelExportSession:
    """
    Keep an Excel workbook in memory for repeated appends, and save on demand.

    This is intended for batching/async saving scenarios to avoid repeatedly calling
    load_workbook() for every single record.
    """

    def __init__(self, *, file_path: str):
        self.file_path = str(file_path)
        self._wb: Workbook | None = None
        self._created_any_sheet: bool = False
        self._dirty: bool = False

    def _open_or_create(self) -> ExportResult:
        if self._wb is not None:
            return ExportResult(ok=True, message="workbook_ready")
        try:
            if os.path.exists(self.file_path):
                wb = load_workbook(self.file_path)
            else:
                wb = Workbook()
        except Exception as e:
            return ExportResult(ok=False, message=f"打开Excel失败: {e}")

        self._wb = wb
        self._created_any_sheet = not (wb.sheetnames == ["Sheet"] and wb.active.title == "Sheet")
        self._dirty = False
        return ExportResult(ok=True, message="workbook_opened")

    def close(self):
        wb = self._wb
        self._wb = None
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

    def append(
        self,
        *,
        save_items: list[str],
        max_points: int,
        sn: str,
        date_text: str,
        analysis_items_data: dict[str, dict[str, Any]],
        analysis_config: dict[str, Any],
        analysis_result_dict: dict[str, tuple[bool, float]],
    ) -> ExportResult:
        ret = self._open_or_create()
        if not ret.ok:
            return ret
        if not isinstance(save_items, list) or len(save_items) == 0:
            return ExportResult(ok=False, message="未选择需要保存的分析项")
        wb = self._wb
        if wb is None:
            return ExportResult(ok=False, message="Excel workbook 未初始化")

        created_any_sheet = self._created_any_sheet

        def ensure_sheet(desired_name: str):
            nonlocal created_any_sheet
            desired_name = _sanitize_sheet_name(desired_name)
            if desired_name in wb.sheetnames:
                return wb[desired_name]
            if not created_any_sheet and wb.sheetnames == ["Sheet"] and wb.active.title == "Sheet":
                wb.remove(wb.active)
            created_any_sheet = True
            return wb.create_sheet(title=desired_name)

        for item_name in save_items:
            item_data = analysis_items_data.get(item_name)
            if not isinstance(item_data, dict):
                continue
            item_type = item_data.get("type")
            if item_type == "Spec":
                # user-invisible: never export
                continue

            if item_type == "AI":
                ws_name = _sanitize_sheet_name(item_name)
                ws = wb[ws_name] if ws_name in wb.sheetnames else ensure_sheet(ws_name)
                header = AI_SHEET_HEADER
                if ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
                    _ensure_header(ws, header)
                label = item_data.get("label") or item_data.get("result") or ""
                ok_score = item_data.get("ok_score")
                ng_score = item_data.get("ng_score")
                model_name = item_data.get("model_name") or item_data.get("model") or ""
                _append_row(ws, [sn, date_text, label, ok_score, ng_score, model_name])
            else:
                result = item_data.get("result")
                xy = _extract_curve_xy(result) if isinstance(result, dict) else None
                if xy is None:
                    continue
                x, y = xy
                x, y = _downsample_xy(x, y, min(int(max_points), EXCEL_MAX_DATA_POINTS))

                ws_base = _sanitize_sheet_name(item_name)
                if ws_base in wb.sheetnames:
                    ws = wb[ws_base]
                    existing = [c.value for c in ws[1][2:] if c.value is not None]
                    expected = [_normalize_header_value(v) for v in x]
                    if not _headers_match(existing, expected):
                        v2_name = _sanitize_sheet_name(f"{item_name}_v2")
                        if v2_name in wb.sheetnames:
                            ws2 = wb[v2_name]
                            existing2 = [c.value for c in ws2[1][2:] if c.value is not None]
                            if _headers_match(existing2, expected):
                                ws = ws2
                            else:
                                v2_name = _make_unique_sheet_name(wb.sheetnames, f"{item_name}_v2")
                                ws = ensure_sheet(v2_name)
                        else:
                            ws = ensure_sheet(v2_name)
                else:
                    ws = ensure_sheet(ws_base)

                header = CURVE_SHEET_HEADER_PREFIX + x
                if ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
                    _ensure_header(ws, header)
                _append_row(ws, [sn, date_text] + y)

            result_tuple = analysis_result_dict.get(item_name)
            if isinstance(result_tuple, tuple) and len(result_tuple) == 2:
                ok_val, deviation = result_tuple
                cfg = analysis_config.get(item_name) if isinstance(analysis_config, dict) else None
                result_sheet_name = _make_margin_sheet_title(item_name)
                rs = wb[result_sheet_name] if result_sheet_name in wb.sheetnames else ensure_sheet(result_sheet_name)
                if rs.max_row < 1 or rs.cell(row=1, column=1).value is None:
                    _ensure_header(rs, RESULT_SHEET_HEADER)

                short_name = _export_short_name(item_name, item_type)
                margin_name = f"{short_name} margin"
                margin_text = _format_margin_value(item_type, deviation)
                unit_text = _export_unit(item_type, cfg if isinstance(cfg, dict) else None)
                result_text = "" if ok_val is None else ("PASS" if bool(ok_val) else "FAIL")
                tol_text, limit_type_text = _export_tolerance(cfg if isinstance(cfg, dict) else None)
                time_text = _format_result_time_minute(date_text)
                _append_row(
                    rs,
                    [sn, time_text, margin_name, margin_text, unit_text, result_text, tol_text, limit_type_text],
                )

        # Ensure at least one sheet exists (openpyxl requires it)
        if len(wb.sheetnames) == 0:
            wb.create_sheet(title="Sheet")

        self._created_any_sheet = created_any_sheet
        self._dirty = True
        return ExportResult(ok=True, message=f"Excel缓存写入: {self.file_path}")

    def save(self) -> ExportResult:
        ret = self._open_or_create()
        if not ret.ok:
            return ret
        wb = self._wb
        if wb is None:
            return ExportResult(ok=False, message="Excel workbook 未初始化")
        if not self._dirty:
            return ExportResult(ok=True, message=f"Excel已是最新: {self.file_path}")

        try:
            wb.save(self.file_path)
            self._dirty = False
            return ExportResult(ok=True, message=f"Excel导出成功: {self.file_path}")
        except PermissionError:
            return ExportResult(ok=False, message=f"Excel文件被占用，请关闭后重试: {self.file_path}")
        except Exception as e:
            return ExportResult(ok=False, message=f"保存Excel失败: {e}")


def _sanitize_sheet_name(name: str) -> str:
    """
    Excel sheet name constraints:
    - max 31 chars
    - cannot contain: []:*?/\\
    """
    if name is None:
        name = ""
    name = str(name).strip()
    name = _INVALID_SHEET_CHARS_RE.sub("_", name)
    if not name:
        name = "Sheet"
    return name[:31]


def _sanitize_filename(name: str) -> str:
    if name is None:
        name = ""
    name = str(name).strip()
    name = _INVALID_FILENAME_CHARS_RE.sub("_", name)
    name = name.strip(". ")
    return name or "analysis_results"


def _make_unique_sheet_name(existing: Iterable[str], desired: str) -> str:
    desired = _sanitize_sheet_name(desired)
    if desired not in existing:
        return desired
    base = desired
    for i in range(2, 1000):
        suffix = f"_{i}"
        truncated = base[: max(0, 31 - len(suffix))]
        candidate = f"{truncated}{suffix}"
        if candidate not in existing:
            return candidate
    # fallback
    return _sanitize_sheet_name(f"{base}_{datetime.now().strftime('%H%M%S')}")


def _headers_match(existing: list[Any], expected: list[Any]) -> bool:
    if len(existing) != len(expected):
        return False
    for a, b in zip(existing, expected):
        if a is None and b is None:
            continue
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            if abs(float(a) - float(b)) > 1e-6:
                return False
        else:
            if str(a) != str(b):
                return False
    return True


def _normalize_header_value(v: Any) -> Any:
    if isinstance(v, (np.floating, float)):
        return round(float(v), 6)
    if isinstance(v, (np.integer, int)):
        return int(v)
    return v


def _downsample_xy(x: list[Any], y: list[Any], max_points: int) -> tuple[list[Any], list[Any]]:
    if max_points <= 0:
        return x, y
    n = min(len(x), len(y))
    x = list(x[:n])
    y = list(y[:n])
    if n <= max_points:
        return x, y
    idx = np.linspace(0, n - 1, num=max_points, dtype=int)
    # ensure strictly increasing indices
    idx = np.unique(idx)
    return [x[int(i)] for i in idx], [y[int(i)] for i in idx]


def _extract_curve_xy(result: dict[str, Any]) -> tuple[list[Any], list[Any]] | None:
    """
    Best-effort extraction of (x,y) from known analysis result schemas.
    """
    if not isinstance(result, dict):
        return None

    candidates = [
        ("frequency_list", "fr_raw"),
        ("frequency_list", "fr"),
        ("frequency_list", "spl_db_raw"),
        ("frequency_list", "spl_db"),
        ("freq_value", "thd_raw"),
        ("freq_value", "thd"),
        ("signal_duration", "signal_spl_raw"),
        ("signal_duration", "signal_spl"),
        ("time_s", "loudness_sone"),
        ("time_s", "sharpness_acum"),
        ("frequency_hz", "pr_db"),
    ]
    for xk, yk in candidates:
        if xk in result and yk in result:
            x = result.get(xk)
            y = result.get(yk)
            if x is None or y is None:
                continue
            if isinstance(x, np.ndarray):
                x = x.tolist()
            if isinstance(y, np.ndarray):
                y = y.tolist()
            if isinstance(x, list) and isinstance(y, list) and len(x) > 0 and len(y) > 0:
                return x, y
    return None


def _ensure_header(ws, header: list[Any]):
    ws.cell(row=1, column=1).value = header[0]
    ws.cell(row=1, column=2).value = header[1]
    for i, v in enumerate(header[2:], start=3):
        ws.cell(row=1, column=i).value = _normalize_header_value(v)


def _append_row(ws, row_values: list[Any]):
    # Ensure header exists; append after the last row
    row_idx = ws.max_row + 1 if ws.max_row >= 1 else 2
    for col_idx, v in enumerate(row_values, start=1):
        ws.cell(row=row_idx, column=col_idx).value = v


def resolve_excel_spool_dir(
    excel_cfg: dict[str, Any],
    *,
    file_path: str | None = None,
    product_model: str | None = None,
) -> str:
    """
    Resolve a per-day spool directory for fast, append-only exporting.

    Rationale: updating a growing .xlsx on every record is fundamentally slow because openpyxl
    must rewrite the whole zip each save. We therefore append rows into CSV spool files during
    production (fast), then build the daily .xlsx in batch (idle / end-of-day).
    """
    xlsx_path = str(file_path or resolve_excel_output_path(excel_cfg, product_model=product_model))
    base, _ext = os.path.splitext(xlsx_path)
    return f"{base}_spool"


def _format_result_time_minute(date_text: str) -> str:
    text = str(date_text or "").strip()
    if not text:
        return ""
    parts = text.split()
    if len(parts) < 2:
        return text
    date_part = parts[0]
    time_part = parts[1]
    t_parts = time_part.split(":")
    if len(t_parts) >= 2:
        return f"{date_part} {t_parts[0]}:{t_parts[1]}"
    return text


def _export_short_name(item_name: str, item_type: Any) -> str:
    """
    Derive a user-facing name for margin output.

    The analysis item name is user-configurable, so we keep the name as-is (no type mapping / parsing).
    """
    raw = str(item_name or "").strip()
    if raw:
        return raw
    fallback = str(item_type or "").strip()
    return fallback or "UNKNOWN"


def _export_unit(item_type: Any, cfg: dict[str, Any] | None) -> str:
    t = str(item_type or "").strip().upper()
    if t in ("HD", "RB"):
        return "%"
    if t == "AI":
        return "%"
    if t in ("SPL", "SPLF", "FR"):
        return "dB"
    if t == "PRB":
        return "phon"
    if t == "LOUD":
        return "sone"
    if t == "SHRP":
        return "acum"
    if t == "ROUGH":
        return "asper"
    if t == "PR":
        return "dB"
    return ""


def _export_tolerance(cfg: dict[str, Any] | None) -> tuple[str, str]:
    """
    Return (tolerance_type, limits_type).

    - import_config/config_dir -> Tolerance curves + Absolute Limits
    - otherwise               -> Absolute Limits + Absolute Limits
    """
    if isinstance(cfg, dict) and (cfg.get("import_config") or cfg.get("config_dir")):
        return "Tolerance curves", "Absolute Limits"
    return "Absolute Limits", "Absolute Limits"


def _format_margin_value(item_type: Any, deviation: Any) -> str:
    if deviation is None:
        return ""
    try:
        v = float(deviation)
    except Exception:
        return str(deviation)

    t = str(item_type or "").strip().upper()
    if t == "AI":
        v *= 100.0
    return f"{v:.2E}"


CURVE_SHEET_HEADER_PREFIX = ["SN", "time"]
AI_SHEET_HEADER = ["SN", "time", "AI result", "OK score(%)", "NG score(%)", "Model"]
RESULT_SHEET_HEADER = ["SN", "time", "Name", "Margin", "Unit", "Result", "Tolerance", "Limit Type"]


def _make_margin_sheet_title(item_name: str) -> str:
    """
    Make an Excel sheet title for the margin sheet, ensuring it ends with " margin".

    Excel sheet title constraints:
    - max 31 chars
    - cannot contain: []:*?/\\
    """
    suffix = " margin"
    base = str(item_name or "").strip()
    base = _INVALID_SHEET_CHARS_RE.sub("_", base)
    if not base:
        base = "Sheet"
    max_base_len = 31 - len(suffix)
    if max_base_len <= 0:
        return _sanitize_sheet_name(suffix.strip() or "margin")
    base = base[:max_base_len]
    return f"{base}{suffix}"[:31]


def _coerce_csv_number(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    text = str(v).strip()
    if text == "":
        return None
    try:
        i = int(text)
        return i
    except Exception:
        pass
    try:
        f = float(text)
        if f.is_integer():
            return int(f)
        return f
    except Exception:
        return text


def _read_csv_header(path: str) -> list[Any] | None:
    # If the CSV is locked by this process (runtime lock), reuse the same handle.
    try:
        header = _RUNTIME_LOCKER.peek_csv_header(path)
        if header:
            return header
    except Exception:
        pass
    try:
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if not header:
                return None
            return list(header)
    except FileNotFoundError:
        return None
    except Exception:
        return None


def _resolve_curve_spool_csv_path(
    spool_dir: str, *, item_name: str, expected_x: list[Any]
) -> tuple[str, str]:
    """
    Return (sheet_name, csv_path) for a curve item. If the existing header mismatches,
    fall back to *_v2 and create a unique name when needed.
    """
    sheet_base = _sanitize_sheet_name(item_name)
    expected = [_normalize_header_value(v) for v in expected_x]

    def _csv_path(sheet_name: str) -> str:
        return os.path.join(spool_dir, f"{_sanitize_filename(sheet_name)}.csv")

    def _header_matches(path: str) -> bool:
        hdr = _read_csv_header(path)
        if not isinstance(hdr, list) or len(hdr) < 3:
            return False
        existing = [_coerce_csv_number(v) for v in hdr[2:]]
        existing = [_normalize_header_value(v) for v in existing]
        return _headers_match(existing, expected)

    p0 = _csv_path(sheet_base)
    if os.path.exists(p0) and _header_matches(p0):
        return sheet_base, p0

    v2_base = _sanitize_sheet_name(f"{item_name}_v2")
    p1 = _csv_path(v2_base)
    if os.path.exists(p1) and _header_matches(p1):
        return v2_base, p1

    if os.path.exists(p0) and not _header_matches(p0):
        try:
            existing_names = [os.path.splitext(f)[0] for f in os.listdir(spool_dir) if f.lower().endswith(".csv")]
        except Exception:
            existing_names = []
        v2_unique = _make_unique_sheet_name(existing_names, v2_base)
        return v2_unique, _csv_path(v2_unique)

    # Either p0 does not exist, or it exists but is empty/bad header; use base.
    return sheet_base, p0


def _append_csv_row(path: str, *, header: list[Any], row: list[Any]) -> ExportResult:
    try:
        new_file = not os.path.exists(path)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            if new_file:
                writer.writerow(header)
            writer.writerow(row)
        return ExportResult(ok=True, message=f"CSV追加成功: {path}")
    except PermissionError:
        return ExportResult(ok=False, message=f"CSV文件被占用，请关闭后重试: {path}")
    except Exception as e:
        return ExportResult(ok=False, message=f"写入CSV失败: {e}")


def _append_csv_row_locked(path: str, *, header: list[Any], row: list[Any]) -> ExportResult:
    return _RUNTIME_LOCKER.append_csv_row(path, header=header, row=row)


def export_analysis_to_csv_spool(
    excel_cfg: dict[str, Any],
    *,
    sn: str,
    date_text: str,
    analysis_items_data: dict[str, dict[str, Any]],
    analysis_config: dict[str, Any],
    analysis_result_dict: dict[str, tuple[bool, float]],
    product_model: str | None = None,
    file_path: str | None = None,
    spool_dir: str | None = None,
) -> ExportResult:
    """
    Fast per-record export: append to CSV spool files instead of rewriting the .xlsx every time.

    Call `build_excel_from_csv_spool()` later to generate the daily .xlsx.
    """
    if not isinstance(excel_cfg, dict) or not excel_cfg.get("enabled", True):
        return ExportResult(ok=True, message="Excel导出未启用")

    save_items = excel_cfg.get("save_items") or []
    if not isinstance(save_items, list) or len(save_items) == 0:
        return ExportResult(ok=False, message="未选择需要保存的分析项")

    try:
        file_path = str(file_path or resolve_excel_output_path(excel_cfg, product_model=product_model))
    except Exception as e:
        return ExportResult(ok=False, message=f"保存目录不可达或无权限: {e}")
    max_points = resolve_excel_max_points(excel_cfg)
    try:
        spool_dir = str(spool_dir or resolve_excel_spool_dir(excel_cfg, file_path=file_path))
        os.makedirs(spool_dir, exist_ok=True)
    except Exception as e:
        return ExportResult(ok=False, message=f"CSV缓存目录不可达或无权限: {e}")

    use_locks = _locks_enabled(excel_cfg) and msvcrt is not None
    append_csv = _append_csv_row_locked if use_locks else _append_csv_row

    # Rotate locks when daily output path changes (long-running app across days).
    if use_locks:
        try:
            prev_xlsx = excel_cfg.get("_runtime_locked_xlsx_path")
            prev_spool = excel_cfg.get("_runtime_locked_spool_dir")
            if prev_xlsx and os.path.normcase(os.path.abspath(str(prev_xlsx))) != os.path.normcase(
                os.path.abspath(str(file_path))
            ):
                _RUNTIME_LOCKER.close_xlsx(str(prev_xlsx))
            if prev_spool and os.path.normcase(os.path.abspath(str(prev_spool))) != os.path.normcase(
                os.path.abspath(str(spool_dir))
            ):
                _RUNTIME_LOCKER.close_csv_under_dir(str(prev_spool))
            excel_cfg["_runtime_locked_xlsx_path"] = str(file_path)
            excel_cfg["_runtime_locked_spool_dir"] = str(spool_dir)
        except Exception:
            pass

    # Lock existing daily XLSX early so operators can't open it and block later rebuilds.
    if use_locks and os.path.exists(file_path):
        try:
            _RUNTIME_LOCKER.get_locked_xlsx(file_path)
        except PermissionError:
            return ExportResult(ok=False, message=f"Excel文件被占用，请关闭后重试: {file_path}")
        except Exception as e:
            return ExportResult(ok=False, message=f"Excel加锁失败: {e}")

    for item_name in save_items:
        item_data = analysis_items_data.get(item_name)
        if not isinstance(item_data, dict):
            continue
        item_type = item_data.get("type")
        if item_type == "Spec":
            continue

        if item_type == "AI":
            sheet_name = _sanitize_sheet_name(item_name)
            csv_path = os.path.join(spool_dir, f"{_sanitize_filename(sheet_name)}.csv")
            header = AI_SHEET_HEADER
            label = item_data.get("label") or item_data.get("result") or ""
            ok_score = item_data.get("ok_score")
            ng_score = item_data.get("ng_score")
            model_name = item_data.get("model_name") or item_data.get("model") or ""
            ret = append_csv(
                csv_path, header=header, row=[sn, date_text, label, ok_score, ng_score, model_name]
            )
            if not ret.ok:
                return ret
        else:
            result = item_data.get("result")
            xy = _extract_curve_xy(result) if isinstance(result, dict) else None
            if xy is None:
                continue
            x, y = xy
            x, y = _downsample_xy(x, y, min(int(max_points), EXCEL_MAX_DATA_POINTS))

            _sheet_name, csv_path = _resolve_curve_spool_csv_path(spool_dir, item_name=item_name, expected_x=x)
            header = CURVE_SHEET_HEADER_PREFIX + [_normalize_header_value(v) for v in x]
            row = [sn, date_text] + list(y)
            ret = append_csv(csv_path, header=header, row=row)
            if not ret.ok:
                return ret

        result_tuple = analysis_result_dict.get(item_name)
        if isinstance(result_tuple, tuple) and len(result_tuple) == 2:
            ok_val, deviation = result_tuple
            cfg = analysis_config.get(item_name) if isinstance(analysis_config, dict) else None
            sheet_name = _make_margin_sheet_title(item_name)
            csv_path = os.path.join(spool_dir, f"{_sanitize_filename(sheet_name)}.csv")

            short_name = _export_short_name(item_name, item_type)
            margin_name = f"{short_name} margin"
            margin_text = _format_margin_value(item_type, deviation)
            unit_text = _export_unit(item_type, cfg if isinstance(cfg, dict) else None)
            result_text = "" if ok_val is None else ("PASS" if bool(ok_val) else "FAIL")
            tol_text, limit_type_text = _export_tolerance(cfg if isinstance(cfg, dict) else None)
            time_text = _format_result_time_minute(date_text)

            ret = append_csv(
                csv_path,
                header=RESULT_SHEET_HEADER,
                row=[sn, time_text, margin_name, margin_text, unit_text, result_text, tol_text, limit_type_text],
            )
            if not ret.ok:
                return ret

    return ExportResult(ok=True, message=f"CSV缓存写入: {spool_dir}")


def build_excel_from_csv_spool(
    excel_cfg: dict[str, Any],
    *,
    file_path: str | None = None,
    spool_dir: str | None = None,
    product_model: str | None = None,
) -> ExportResult:
    """
    Build (or rebuild) the daily .xlsx from CSV spool files using openpyxl write_only mode.

    This is intended to run off the critical production path (idle / end-of-day / background).
    """
    if not isinstance(excel_cfg, dict) or not excel_cfg.get("enabled", True):
        return ExportResult(ok=True, message="Excel导出未启用")

    use_locks = _locks_enabled(excel_cfg) and msvcrt is not None

    try:
        xlsx_path = str(file_path or resolve_excel_output_path(excel_cfg, product_model=product_model))
        spool_path = str(
            spool_dir
            or resolve_excel_spool_dir(excel_cfg, file_path=xlsx_path, product_model=product_model)
        )
    except Exception as e:
        return ExportResult(ok=False, message=f"保存目录不可达或无权限: {e}")
    if not os.path.isdir(spool_path):
        return ExportResult(ok=True, message="未找到CSV缓存目录，跳过生成Excel")

    try:
        csv_files = sorted([f for f in os.listdir(spool_path) if f.lower().endswith(".csv")])
    except Exception as e:
        return ExportResult(ok=False, message=f"读取CSV缓存目录失败: {e}")

    if not csv_files:
        return ExportResult(ok=True, message="CSV缓存为空，跳过生成Excel")

    # Lock existing xlsx so operators can't open it while program is running.
    if use_locks and os.path.exists(xlsx_path):
        try:
            _RUNTIME_LOCKER.get_locked_xlsx(xlsx_path)
        except PermissionError:
            return ExportResult(ok=False, message=f"Excel文件被占用，请关闭后重试: {xlsx_path}")
        except Exception as e:
            return ExportResult(ok=False, message=f"Excel加锁失败: {e}")

    # Quick skip if xlsx is already newer than all spool csv files.
    try:
        latest_csv_mtime = max(os.path.getmtime(os.path.join(spool_path, f)) for f in csv_files)
        if os.path.exists(xlsx_path) and os.path.getmtime(xlsx_path) >= latest_csv_mtime:
            return ExportResult(ok=True, message=f"Excel宸叉槸鏈€鏂? {xlsx_path}")
    except Exception:
        pass

    wb = Workbook(write_only=True)

    for csv_name in csv_files:
        sheet_title = _sanitize_sheet_name(os.path.splitext(csv_name)[0])
        ws = wb.create_sheet(title=sheet_title)
        csv_path = os.path.join(spool_path, csv_name)
        try:
            if use_locks:
                r, _hdr, content = _RUNTIME_LOCKER.read_csv_snapshot(csv_path)
                if not r.ok:
                    return ExportResult(ok=False, message=r.message)
                reader = csv.reader(io.StringIO(content))
            else:
                f = open(csv_path, "r", newline="", encoding="utf-8-sig")
                reader = csv.reader(f)

            try:
                header = next(reader, None)
                if not header:
                    continue

                # Detect sheet type for lightweight parsing (keep first 2 columns as text)
                is_ai = len(header) >= 6 and ("AI" in str(header[2]))
                is_margin_result = (
                    len(header) >= 6
                    and str(header[2]).strip().lower() == "name"
                    and str(header[3]).strip().lower() == "margin"
                )
                is_legacy_result = len(header) == 4 and ("结果" in str(header[2]) or "偏差" in str(header[3]))

                ws.append(list(header))

                if not is_ai and not is_margin_result and not is_legacy_result:
                    # Curve-like sheet: parse numeric cells (col>=3) to numbers.
                    for row in reader:
                        if not row or len(row) < len(header):
                            continue
                        padded = list(row[: len(header)])
                        out = [padded[0], padded[1]]
                        for v in padded[2:]:
                            out.append(_coerce_csv_number(v))
                        ws.append(out)
                elif is_ai:
                    for row in reader:
                        if not row or len(row) < len(header):
                            continue
                        padded = list(row[: len(header)])
                        out = [
                            padded[0],
                            padded[1],
                            padded[2],
                            _coerce_csv_number(padded[3]),
                            _coerce_csv_number(padded[4]),
                            padded[5] if len(padded) > 5 else None,
                        ]
                        ws.append(out)
                elif is_margin_result:
                    # Keep all fields as text to preserve formats like "6.00E-01".
                    for row in reader:
                        if not row or len(row) < len(header):
                            continue
                        ws.append(list(row[: len(header)]))
                else:
                    # legacy result sheet
                    for row in reader:
                        if not row or len(row) < len(header):
                            continue
                        padded = list(row[: len(header)])
                        out = [padded[0], padded[1], padded[2], _coerce_csv_number(padded[3])]
                        ws.append(out)
            finally:
                if not use_locks:
                    try:
                        f.close()
                    except Exception:
                        pass
        except Exception as e:
            return ExportResult(ok=False, message=f"读取CSV失败: {csv_path}: {e}")

    if len(wb.sheetnames) == 0:
        wb.create_sheet(title="Sheet")

    if use_locks:
        return _RUNTIME_LOCKER.save_xlsx(xlsx_path, wb)

    try:
        wb.save(xlsx_path)
        return ExportResult(ok=True, message=f"Excel生成成功: {xlsx_path}")
    except PermissionError:
        return ExportResult(ok=False, message=f"Excel文件被占用，请关闭后重试: {xlsx_path}")
    except Exception as e:
        return ExportResult(ok=False, message=f"保存Excel失败: {e}")


def export_analysis_to_excel(
    excel_cfg: dict[str, Any],
    *,
    sn: str,
    date_text: str,
    analysis_items_data: dict[str, dict[str, Any]],
    analysis_config: dict[str, Any],
    analysis_result_dict: dict[str, tuple[bool, float]],
    product_model: str | None = None,
    file_path: str | None = None,
) -> ExportResult:
    if not isinstance(excel_cfg, dict) or not excel_cfg.get("enabled", True):
        return ExportResult(ok=True, message="Excel导出未启用")

    save_items = excel_cfg.get("save_items") or []
    if not isinstance(save_items, list) or len(save_items) == 0:
        return ExportResult(ok=False, message="未选择需要保存的分析项")

    try:
        file_path = str(file_path or resolve_excel_output_path(excel_cfg, product_model=product_model))
    except Exception as e:
        return ExportResult(ok=False, message=f"保存目录不可达或无权限: {e}")
    max_points = resolve_excel_max_points(excel_cfg)
    use_locks = _locks_enabled(excel_cfg) and msvcrt is not None

    locked_xlsx = None
    if use_locks:
        try:
            locked_xlsx = _RUNTIME_LOCKER.get_locked_xlsx(file_path)
        except PermissionError:
            return ExportResult(ok=False, message=f"Excel文件被占用，请关闭后重试: {file_path}")
        except Exception as e:
            return ExportResult(ok=False, message=f"Excel加锁失败: {e}")

    # Load or create workbook
    try:
        if locked_xlsx is not None:
            with locked_xlsx.lock:
                locked_xlsx.f.seek(0, os.SEEK_END)
                size = int(locked_xlsx.f.tell())
                locked_xlsx.f.seek(0)
                wb = load_workbook(locked_xlsx.f) if size > 0 else Workbook()
        else:
            wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
    except Exception as e:
        return ExportResult(ok=False, message=f"打开Excel失败: {e}")

    # If it's a new workbook, remove default sheet when we create our first real sheet
    created_any_sheet = False

    def ensure_sheet(desired_name: str):
        nonlocal created_any_sheet
        desired_name = _sanitize_sheet_name(desired_name)
        if desired_name in wb.sheetnames:
            return wb[desired_name]
        if not created_any_sheet and wb.sheetnames == ["Sheet"] and wb.active.title == "Sheet":
            wb.remove(wb.active)
        created_any_sheet = True
        return wb.create_sheet(title=desired_name)

    for item_name in save_items:
        item_data = analysis_items_data.get(item_name)
        if not isinstance(item_data, dict):
            continue
        item_type = item_data.get("type")
        if item_type == "Spec":
            # user-invisible: never export
            continue

        if item_type == "AI":
            ws_name = _sanitize_sheet_name(item_name)
            ws = wb[ws_name] if ws_name in wb.sheetnames else ensure_sheet(ws_name)
            header = AI_SHEET_HEADER
            if ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
                _ensure_header(ws, header)
            label = item_data.get("label") or item_data.get("result") or ""
            ok_score = item_data.get("ok_score")
            ng_score = item_data.get("ng_score")
            model_name = item_data.get("model_name") or item_data.get("model") or ""
            _append_row(ws, [sn, date_text, label, ok_score, ng_score, model_name])
        else:
            result = item_data.get("result")
            xy = _extract_curve_xy(result) if isinstance(result, dict) else None
            if xy is None:
                continue
            x, y = xy
            x, y = _downsample_xy(x, y, min(max_points, EXCEL_MAX_DATA_POINTS))

            ws_base = _sanitize_sheet_name(item_name)
            if ws_base in wb.sheetnames:
                ws = wb[ws_base]
                existing = [c.value for c in ws[1][2:] if c.value is not None]
                expected = [_normalize_header_value(v) for v in x]
                if not _headers_match(existing, expected):
                    v2_name = _sanitize_sheet_name(f"{item_name}_v2")
                    if v2_name in wb.sheetnames:
                        ws2 = wb[v2_name]
                        existing2 = [c.value for c in ws2[1][2:] if c.value is not None]
                        if _headers_match(existing2, expected):
                            ws = ws2
                        else:
                            v2_name = _make_unique_sheet_name(wb.sheetnames, f"{item_name}_v2")
                            ws = ensure_sheet(v2_name)
                    else:
                        ws = ensure_sheet(v2_name)
            else:
                ws = ensure_sheet(ws_base)

            header = CURVE_SHEET_HEADER_PREFIX + x
            if ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
                _ensure_header(ws, header)
            _append_row(ws, [sn, date_text] + y)

        result_tuple = analysis_result_dict.get(item_name)
        if isinstance(result_tuple, tuple) and len(result_tuple) == 2:
            ok_val, deviation = result_tuple
            cfg = analysis_config.get(item_name) if isinstance(analysis_config, dict) else None
            result_sheet_name = _make_margin_sheet_title(item_name)
            rs = wb[result_sheet_name] if result_sheet_name in wb.sheetnames else ensure_sheet(result_sheet_name)
            if rs.max_row < 1 or rs.cell(row=1, column=1).value is None:
                _ensure_header(rs, RESULT_SHEET_HEADER)

            short_name = _export_short_name(item_name, item_type)
            margin_name = f"{short_name} margin"
            margin_text = _format_margin_value(item_type, deviation)
            unit_text = _export_unit(item_type, cfg if isinstance(cfg, dict) else None)
            result_text = "" if ok_val is None else ("PASS" if bool(ok_val) else "FAIL")
            tol_text, limit_type_text = _export_tolerance(cfg if isinstance(cfg, dict) else None)
            time_text = _format_result_time_minute(date_text)
            _append_row(
                rs,
                [sn, time_text, margin_name, margin_text, unit_text, result_text, tol_text, limit_type_text],
            )

    # Ensure at least one sheet exists (openpyxl requires it)
    if len(wb.sheetnames) == 0:
        wb.create_sheet(title="Sheet")

    # Save workbook (avoid rename/replace; some environments may block file renames)
    if locked_xlsx is not None:
        try:
            with locked_xlsx.lock:
                locked_xlsx.f.seek(0)
                locked_xlsx.f.truncate(0)
                wb.save(locked_xlsx.f)
                try:
                    locked_xlsx.f.flush()
                except Exception:
                    pass
            return ExportResult(ok=True, message=f"Excel导出成功: {file_path}")
        except PermissionError:
            return ExportResult(ok=False, message=f"Excel文件被占用，请关闭后重试: {file_path}")
        except Exception as e:
            return ExportResult(ok=False, message=f"保存Excel失败: {e}")

    try:
        wb.save(file_path)
        return ExportResult(ok=True, message=f"Excel导出成功: {file_path}")
    except PermissionError:
        return ExportResult(ok=False, message=f"Excel文件被占用，请关闭后重试: {file_path}")
    except Exception as e:
        return ExportResult(ok=False, message=f"保存Excel失败: {e}")
